"""Excel document parser using openpyxl."""

from __future__ import annotations

import io
import time
from typing import Any

import structlog

from app.docprocessor.base import ExtractedSection, ExtractedTable, ExtractionResult

logger = structlog.stdlib.get_logger()


class ExcelParser:
    """Extract structured data from Excel files using openpyxl."""

    MAX_SHEETS = 50
    MAX_ROWS_PER_SHEET = 10_000
    MAX_FILE_SIZE = 100_000_000  # 100 MB
    MAX_TOTAL_CELLS = 1_000_000  # Guard against decompression bombs

    async def parse(
        self,
        file_path: str | None = None,
        file_bytes: bytes | None = None,
        filename: str = "",
    ) -> ExtractionResult:
        """Parse Excel file and extract structured content."""
        start = time.monotonic()
        result = ExtractionResult(source_type="excel", source_name=filename or file_path or "unknown.xlsx")

        if file_bytes and len(file_bytes) > self.MAX_FILE_SIZE:
            result.metadata["error"] = f"File too large ({len(file_bytes)} bytes, max {self.MAX_FILE_SIZE})"
            result.extraction_duration_ms = int((time.monotonic() - start) * 1000)
            return result

        try:
            import openpyxl
        except ImportError:
            logger.error("openpyxl_not_installed")
            result.metadata["error"] = "openpyxl is not installed"
            result.extraction_duration_ms = int((time.monotonic() - start) * 1000)
            return result

        try:
            if file_path:
                wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
            else:
                wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)  # type: ignore[arg-type]

            try:
                result.metadata["sheet_count"] = len(wb.sheetnames)
                all_text_parts: list[str] = []
                cells_read: list[int] = [0]

                for sheet_name in wb.sheetnames[: self.MAX_SHEETS]:
                    if cells_read[0] >= self.MAX_TOTAL_CELLS:
                        break
                    ws = wb[sheet_name]
                    try:
                        section, table = self._extract_sheet(ws, sheet_name, cells_read)
                        if section:
                            result.sections.append(section)
                            all_text_parts.append(section.content)
                        if table:
                            result.tables.append(table)
                    except Exception as exc:
                        logger.warning("excel_sheet_extraction_failed", sheet=sheet_name, error=str(exc))

                result.raw_text = "\n\n".join(all_text_parts)
            finally:
                wb.close()

        except Exception as exc:
            logger.error("excel_extraction_failed", error=str(exc), filename=filename)
            result.metadata["error"] = str(exc)

        result.extraction_duration_ms = int((time.monotonic() - start) * 1000)
        return result

    def _extract_sheet(
        self,
        ws: Any,
        sheet_name: str,
        cells_read: list[int] | None = None,
    ) -> tuple[ExtractedSection | None, ExtractedTable | None]:
        """Extract content from a single worksheet."""
        rows_data: list[list[str]] = []
        if cells_read is None:
            cells_read = [0]

        for row_count, row in enumerate(ws.iter_rows()):
            if row_count >= self.MAX_ROWS_PER_SHEET:
                break
            if cells_read[0] >= self.MAX_TOTAL_CELLS:
                logger.warning("excel_cell_limit_reached", sheet=sheet_name, cells=cells_read[0])
                break
            row_values = []
            for cell in row:
                cells_read[0] += 1
                value = cell.value
                if value is None:
                    row_values.append("")
                else:
                    row_values.append(str(value).strip())
            # Skip completely empty rows
            if any(v for v in row_values):
                rows_data.append(row_values)

        if not rows_data:
            return None, None

        # Detect header row: first row where all non-empty cells are strings
        headers = rows_data[0]
        data_rows = rows_data[1:] if len(rows_data) > 1 else []

        table = ExtractedTable(
            headers=headers,
            rows=data_rows,
        )

        # Build text representation for RAG
        text_lines = [f"Sheet: {sheet_name}"]
        text_lines.append(" | ".join(headers))
        text_lines.append("-" * 40)
        for row in data_rows[:50]:  # Cap text representation at 50 rows
            text_lines.append(" | ".join(row))
        if len(data_rows) > 50:
            text_lines.append(f"... and {len(data_rows) - 50} more rows")

        section = ExtractedSection(
            title=sheet_name,
            content="\n".join(text_lines),
            level=1,
            tables=[table],
        )

        return section, table
